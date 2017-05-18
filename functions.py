# ******************************************************************************
# Author: Son D. Ngo
# Filename: functions.py
# Contributor: Steve Papaccio
# Updated: 5/17/2017

# Description: Script that implements functions that will be used for 
# 		interface.py, including reading files, running the conform and merging
#		steps. This file can be understood as the backend of program.
# ******************************************************************************

# Modules & Packages
import os
import re
import csv
import xlrd
import xlsxwriter
import datetime
import configparser
import sys
import tkinter as tk
from tqdm import tqdm
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# global constants
def CONFORM_FLAG():
	return 0

def MERGE_FLAG():
	return 1

def FULL_NAME_FLAG():
	return 0

def ABBR_NAME_FLAG():
	return 1

# global variables for data structures
dictionary_input_csv = ""
user_input_csv = ""
dictionary_input = ""
file_name = ""
non_match_file_name = ""
file_format = ".xlsx"
local_path = ""
columns = defaultdict(list)
dictionary = defaultdict(list)
full_mapper = {};
no_match = [];
match = [];
header = [];

# UTILITY FUNCTIONS 
# convert excel file to csv file 
def csv_from_excel(inFile):
	wb = xlrd.open_workbook(inFile)
	sh = wb.sheet_by_index(0)
	outFile = inFile[0 : inFile.find((".xlsx", ".xls")[inFile.endswith(".xlsx") != -1])] + '_TEMP_DELETE.csv';
	your_csv_file = open(outFile, 'w')
	wr = csv.writer(your_csv_file, quoting = csv.QUOTE_ALL)

	for rownum in range(sh.nrows):
                try:
                        wr.writerow(sh.row_values(rownum))                        
                except ValueError:
                        print("Error in dictionary on row: ")
                        print(rownum)

	your_csv_file.close()

	return outFile

# convert xls to xlsx
def xlsx_from_xls(inFile):
    book_xls = xlrd.open_workbook(inFile)
    book_xlsx = Workbook()
    outFile = inFile[0 : inFile.find(".xls")] + ".xlsx"

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(outFile)

    return outFile

#write conform/non_match result to file as xlsx
def write_conform_result(data, header, outFile):
	book = xlsxwriter.Workbook(outFile)
	sh = book.add_worksheet()

	# if writing output file
	# write out the title first, and then each entry in data is a row
	if header:
		c = 0;
		for title in header:
			sh.write(0, c, title)
			r = 1
			for value in data[title]:
				sh.write(r, c, value)
				r += 1
			c += 1

	# if writing ummatched file or dictionary
	# three indicated columns, and each row is the unmatched name
	else:
		sh.write(0, 0, "AS ENTERED")
		sh.write(0, 1, "FULL NAME")
		sh.write(0, 2, "ABBREVIATED NAME")
		r = 1;
		for entry in data:
			sh.write(r, 0, entry)
			r += 1

	book.close()

# write update result to file as xlsx
def write_update_result(data, outFile):
	book = xlsxwriter.Workbook(outFile)
	sh = book.add_worksheet()

	# only two columns, and each row is the dictionary key
	sh.write(0, 0, "AS ENTERED")
	sh.write(0, 1, "FULL NAME")
	sh.write(0, 2, "ABBREVIATED NAME")
	r = 1;
	for entry in data:
		sh.write(r, 0, entry)
		r += 1

	book.close()

# get header (i.e columns from input file)
def getHeader():
	global header
	return header

# return the path to non_match file
def getNonMatchFilePath():
	global local_path, non_match_file_name 
	return local_path + "/" + non_match_file_name + file_format

# return the path to conformed file
def getConformedFilePath():
	global local_path, conformed_file_name
	return local_path + "/" + conformed_file_name + file_format

# return the path to data/output folder
def getOutputFolderPath():
	global local_path
	return local_path

# delete all temporary file
def clean_temp_file():
	global local_path

	outputDir = os.getcwd()
	if outputDir:
		for f in os.listdir(outputDir):
			if ("_TEMP_DELETE") in f:
				os.remove(os.path.join(outputDir, f))

	if local_path:
		for f in os.listdir(local_path):
			if ("_TEMP_DELETE") in f:
				os.remove(os.path.join(local_path, f))

# read from input file 
def readInputFile(userInput, flag):

	global header, file_name, user_input, user_input_csv, local_path, columns	

	# clear previous data in memory
	if columns:
		columns.clear()

	user_input = userInput

	# converting steps
	if (user_input.endswith(".xlsx") or user_input.endswith(".xls")):
		user_input_csv = csv_from_excel(user_input)
	elif (user_input.endswith(".csv")):
		user_input_csv = userInput
	else:
		messagebox.showerror("Error", "Unsupported file extension. Please choose .xlsx, .xls or .csv files!")
		return False

	print("reading in " + userInput)

	# set up the user input map
	try:
		with open(user_input_csv) as f:
			reader = csv.DictReader(f)
			header = reader.fieldnames
			for row in reader:
				for (k,v) in row.items():
					columns[k].append(v)
	except IOError:
		messagebox.showerror(
			"Open file",
			"Cannot open this file\n(%s)" % dictionary_input_csv
		)
		return False

	# either update conform file or non-match file depending
	# on the action flag
	if flag == CONFORM_FLAG():
		file_name = user_input[user_input.rfind("/")+1:user_input.rfind(".")]
	else:
		non_match_file_name = user_input[user_input.rfind("/")+1:user_input.rfind(".")]

	local_path = user_input[0:user_input.rfind("/")]

	return True

# read from dictionary file
def readDictionaryFile(dictInput, flag):

	global dictionary_input_csv, dictionary_input, dictionary, columns, full_mapper, abbr_mapper

	# clear previous data in memory
	if dictionary:
		dictionary.clear()

	dictionary_input = dictInput

	# conversion steps
	if (dictInput.endswith(".xlsx") or dictInput.endswith(".xls")):
		dictionary_input_csv = csv_from_excel(dictInput)
	elif (dictInput.endswith(".csv")):
		dictionary_input_csv = dictInput
	else:
		messagebox.showerror("Error", "Unsupported file extension. \
			Please choose .xlsx, .xls or .csv files!")
		return False

	print("Reading dictionary " + dictionary_input_csv)		

	# set up the dictionary map
	try:
		with open(dictionary_input_csv) as f:
			reader = csv.DictReader(f)
			for row in reader:
				for (k,v) in row.items():
					dictionary[k].append(v)
	except IOError:
		messagebox.showerror(
			"Open file",
			"Cannot open this file\n(%s)" % dictionary_input_csv
		)
		return False

	duplicateMapper = {}
	full_mapper = {}
	abbr_mapper = {}

	# create full_mapper and abbr_mapper using the dictionary
	for index in range(0,len(dictionary["AS ENTERED"])):
		originalKey = (dictionary["AS ENTERED"])[index]
		key = originalKey.lower()
		key = key.strip();
		key = " ".join(key.split())
		full_value = (dictionary["FULL NAME"])[index]
		abbr_value = (dictionary["ABBREVIATED NAME"])[index]

		# detect that there are conflicts of keys in the dictionary. Alert the user with
		# all the values of the first-found duplicated key
		if key in full_mapper.keys():
			print("Duplicated keys in dictionary: ")
			for item in duplicateMapper[key]:
				print(item)
			print(originalKey)
			messagebox.showerror("Error!", "Found duplicate in dictionary. \
				Refer to the terminal window for more information. \
				Please clean the dictionary to proceed")
			return False
		
		if not key in duplicateMapper:
			duplicateMapper[key] = [originalKey]
		else:
			duplicateMapper[key].push(originalKey)

		full_mapper[key] = full_value
		abbr_mapper[key] = abbr_value

	return True

# conform functions
def conform(column_name, name_flag):	

	global file_name, non_match_file_name, conformed_file_name, dictionary_input_csv, \
		user_input_csv, columns, full_mapper, abbr_mapper

	# recipients is a list of all the pre-conformed names
	recipients = columns[column_name]	
	
	non_match_count = 0
	match_count = 0
	index = 0
	no_match = []
	match = []
	curr_mapper = {}	

	# choose which mapper to use: full name or abbreviated name
	if name_flag == FULL_NAME_FLAG():
		curr_mapper = full_mapper
	else:
		curr_mapper = abbr_mapper

	# conform by going through each name to be conformed, and check
	# if that name is in the dictionary (exact match).
	for item in tqdm(recipients, desc="CONFORMING: "):

		# remove trailing, leading and multiple-in-text white spaces
		# then turn the name to lower case
		item = item.strip();
		item = " ".join(item.split())
		processedItem = item.lower();

		# FUTURE WORK: preprocess special symbols???

		# if the item has a matching in the mapper, then perform
		# the match. 
		if processedItem in curr_mapper.keys():
			match_count += 1
			
			# only record unique values in the match list
			if not processedItem in match:
				match.append(processedItem)

			# only conform the current name is there is a mapping,
			# otherwise keep the current name.
			if curr_mapper[processedItem]:
				columns[column_name][index] = curr_mapper[processedItem]

		# there is no match, keep the original.
		else:
			non_match_count += 1
			if not item in no_match:
				no_match.append(item)

		index += 1

	# producing results
	conformed_file_name = file_name + "_conformed"
	non_match_file_name = file_name + "_non_matched"

	print("\n=============BEGIN WRITING=============")
 
	print("Writing conformed result to:")
	print(conformed_file_name + file_format)
	write_conform_result(columns, header, local_path + "/" + conformed_file_name + file_format)

	print("Writing non_matched result to:")
	print(non_match_file_name + file_format)
	write_conform_result(no_match, [], local_path + "/" + non_match_file_name + file_format)

	print("=============BEGIN CONFORMING=============")

	# prints
	print("\n=============SUMMARY RESULT=============")
	print("Length of dictionary: " + str(len(curr_mapper)))
	print("Number of rows processed: " + str(len(recipients)))
	print("Number of non_match (including duplicates): " + str(non_match_count))
	print("Number of non_match (excluding duplicates): " + str(len(no_match)))
	print("Number of matches (including duplicates): " + str(match_count))
	print("Number of matches (excluding duplicates): " + str(len(match)))
	print("==========================================\n")

	print("\nConforming Process Finished.\n")

	# remove intermediary files
	if (user_input_csv != "" and "_TEMP_DELETE" in user_input_csv):
		os.remove(user_input_csv)
		user_input_csv = ""

	if (dictionary_input_csv != "" and "_TEMP_DELETE" in user_input_csv):
		os.remove(dictionary_input_csv)
		dictionary_input_csv = ""

# merge functionality
def merge():
	global dictionary_input, user_input, dictionary_input_csv, user_input_csv, full_mapper

	book = load_workbook(dictionary_input)
	sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
	sheet.cell(row = 1, column = 4).value = "DATE UPDATED"

	index = 0;
	removedNonMatch = [];

	print("\n=============BEGIN MERGING=============")

	# go through each item in the updated non_matched file, and
	# append the match into the dictionary (if there is any). 
	# Then, the item will be removed from the non_matched file.
	# If an abbreviation is also specified, then update this as
	# well in the dictionary. If not, the abbreviated name will
	# be empty in the dictionary.
	for item in tqdm(columns["AS ENTERED"], desc="MERGING: "):
		conformed_name = columns["FULL NAME"][index]
		abbreviated_name = columns["ABBREVIATED NAME"][index]

		# remove trailing, leading and multiple-in-text white spaces
		# then turn the name to lower case
		item = item.strip();
		item = " ".join(item.split())
		processedItem = item.lower();

		# if the updated conformed name is not found, then move on
		if not conformed_name:
			# if removedNonMatch:
			# 	savedPath = dictionary_input[0:dictionary_input.find(".xlsx")] + "_updated.xlsx"
			# 	print("Save updated dictionary to: " + savedPath)
			# 	book.save(savedPath)
			# break
			index += 1
			continue

		# FUTURE WORK: Resolve when there are two same keys (in the AS_ENTERED column)
		# elif processedItem in full_mapper.keys():
			# if conformed_name != full_mapper[processedItem]:
			# 	print("Found a conflict between user and dictionary")
			# 	print("AS ENTERED: " + item)
			# 	print("FULL NAME (user): " + conformed_name)
			# 	print("FULL NAME (dictionary): " + full_mapper[processedItem])
			# 	print("Which one would you like to keep? (0 for user, 1 for dictionary, or type a new name")
			# 	message = input()
			# 	if message == "1":
			# 		conformed_name = full_mapper[processedItem]
			# 	elif message != "0":
			# 		conformed_name = message

				# implement using message box instead, think about this!
				# if messagebox.askokcancel("Dictionary Conflict!", message):
				# 	conformed_name = full_mapper[processedItem]
				# elif:
				# 	conformed_name = message
		
		# add to temporary worksheet, with the new conformed name. For the abbreviated name,
		# if the user doesn't specify, then leave the field empty
		removedNonMatch.append(item)
		row_to_write = sheet.max_row + 1
		sheet.cell(row = row_to_write, column = 1).value = item
		sheet.cell(row = row_to_write, column = 2).value = conformed_name
		sheet.cell(row = row_to_write, column = 3).value = abbreviated_name if abbreviated_name else ""
		sheet.cell(row = row_to_write, column = 4).value = datetime.date.today()
		print("Conform " + "\"" + item + "\"" + " to:")
		print(conformed_name)
		index += 1

	print("=============FINISH MERGING=============")

	print("Number of non-match removed: " + str(len(removedNonMatch)))

	print("\n=============BEGIN WRITING=============")

	# write merged dictionary to a new file if there is an update
	if removedNonMatch:
		savedPath = dictionary_input[0:dictionary_input.find(".xlsx")] + "_old_" + datetime.datetime.today().strftime('%m-%d-%Y-%H-%M-%S') + ".xlsx"
		os.rename(dictionary_input, savedPath);
		print("Writing updated dictionary to:")		
		book.save(dictionary_input)

		# update non-match file (remove the ones that have been updated)
		columns["AS ENTERED"] = set(columns["AS ENTERED"]) - set(removedNonMatch)
		os.remove(user_input)
		print("updating " + user_input)
		write_update_result(columns["AS ENTERED"], user_input)

	print("=============FINISH WRITING WRITING=============")

	# remove intermediary file
	if (user_input_csv != "" and "_TEMP_DELETE" in user_input_csv):
		os.remove(user_input_csv)
		user_input_csv = ""

	if (dictionary_input_csv != "" and "_TEMP_DELETE" in user_input_csv):
		os.remove(dictionary_input_csv)
		dictionary_input_csv = ""

	print("\nMerging Process Finished.\n")

